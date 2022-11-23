import sys

from scipy.fftpack import fft, hilbert
from obspy.signal.trigger import classic_sta_lta, recursive_sta_lta, carl_sta_trig, delayed_sta_lta
from scipy import signal
import matplotlib as mpl
import numpy as np
import matplotlib.pyplot as plt
import openpyxl




#
mpl.rcParams['font.sans-serif'] = ['SimHei']  # 显示中文
mpl.rcParams['axes.unicode_minus'] = False  # 显示负号


# data = openpyxl.load_workbook('D:\python_code\\2022.08.14现场测量数据\\2022.8.14-1-3.xlsx')
data = openpyxl.load_workbook('D:\\STUDY\lab\\freshman\\2022.09.02 现场试验\\2022.9.2测量1.xlsx')
# data1 = openpyxl.load_workbook('D:\python_code\\2022.08.14现场测量数据\\2022.8.14-1-3.xlsx')


ws = data.active # 得到当前活跃表单的对象（打开该xlsx文件，直接看到的表单就为活跃表单）
# ws1 = data1.active # 得到当前活跃表单的对象（打开该xlsx文件，直接看到的表单就为活跃表单）

channel1_data = []
# channel2_data = []
# 读取方式1：遍历每一个有数据的行
for line in ws:
    channel1 = line[0].value
    # channel2 = line[1].value
    channel1_data.append(channel1)
    # channel2_data.append(channel2)

channel1_data = np.array(channel1_data[1:])
# channel2_data = np.array(channel2_data[1:])


#截取

min_index = np.argmin(channel1_data)
channel1_data = channel1_data[min_index:]
# channel2_data = channel2_data[min_index:]
channel1_data = channel1_data-np.mean(channel1_data)
# channel2_data = channel2_data-np.mean(channel2_data)


# plt.figure(figsize=(20, 10),facecolor='w')
# plt.plot(channel1_data,c= "r")
# plt.xlabel('采样点数')
# plt.title("第二次",fontsize=15)
# plt.grid(True)
# plt.tight_layout()
# plt.show()


 #
###################################带通滤波############################################

#采样率：250
def butter_filter(data,a,b):
    b, a = signal.butter(2, [a,b], 'bandpass')  # 配置滤波器 8 表示滤波器的阶数
    filtedData = signal.filtfilt(b, a, data)  # data为要过滤的信号


    # plt.figure(figsize=(20, 10),facecolor='w')
    # plt.plot(filtedData)
    # # plt.xticks(range(0,len(filtedData)))
    #
    # # plt.tick_params(labelsize=30)
    # plt.xlabel("采样点数")
    # plt.title('信号带通滤波')
    # plt.grid(True)
    # plt.show()
    return filtedData
filtedData_0_10 = butter_filter(channel1_data,0.008,0.08)    #0-10hz

filtedData_10_20 = butter_filter(channel1_data,0.08,0.16)   #10-20hz


# # # ###################################快速傅里叶变换############################################
def data_fft(y,Fs):

    fft_y = fft(y)  # 快速傅里叶变换

    N = len(y)
    x = np.arange(N)  # 频率个数
    half_x = x[range(int(N / 2))]  # 取一半区间

    abs_y = np.abs(fft_y)  # 取复数的绝对值，即复数的模(双边频谱)
    angle_y = np.angle(fft_y)  # 取复数的角度
    normalization_y = abs_y / N  # 归一化处理（双边频谱）
    normalization_half_y = normalization_y[range(int(N / 2))]  # 由于对称性，只取一半区间（单边频谱）
    normalization_half_x = half_x * (Fs / N)

    # plt.plot(x[:2000], y[0:2000])
    # plt.plot(x, y)
    #
    # plt.title('滤波后波形')
    # plt.grid(b=True)
    # plt.show()

    # print("主频",normalization_half_x[np.argmax(normalization_half_y[10:])+10])
    # print(normalization_half_y.shape)
    # print(normalization_half_x.shape)
    # print(normalization_half_y.shape)
    # plt.figure(figsize=(20, 10),facecolor='w')
    # plt.plot(normalization_half_x, normalization_half_y, 'blue')
    # # plt.xlim(0,100)
    # plt.xlabel('F (Hz)')
    # plt.title('频谱图', fontsize=15)
    # plt.grid(True)
    # plt.show()
    return normalization_half_x[np.argmax(normalization_half_y[10:])+10]

# data_fft(datarec,framerate)
main_freq = data_fft(channel1_data[20:500],250)

v0 = round(2*main_freq*9.6,2)
# print(v0)


###############(速度)深度校正#######################################
def cal_depth(v0, times):
    depth = 0
    interval_depth = 0
    speed = v0  #初始速度
    sampling_interval = 0.004  #采样时间间隔
    depth_interval_size = 33   ##深度间隔
    speed_increment = 0.6     #速度增量

    num_time_period = int(times / sampling_interval)

    for i in range(num_time_period):

        if interval_depth >= depth_interval_size :
            speed += speed_increment
            interval_depth = 0
        else:
            interval_depth += speed * sampling_interval

        depth += speed * sampling_interval

    return depth

#####################################################STA/LTA#######################################
framerate = 250
######  https://voidk2.github.io/2019/05/17/sgylog/
#预估深度 ： Predicted depth
def STA_LTA(data,nsta,nlta,a,predicted_depth):

    import matplotlib.pyplot as plt
    n = int(2*predicted_depth*framerate/v0)
    start = int(2 * (predicted_depth - 800) * framerate / v0)
    data = data[:n]
    npts =len(data)  ##采样点个数
    redata = np.square(data)

    cft = recursive_sta_lta(redata, nsta, nlta)  #递归STA/LTA


    t = np.arange(npts, dtype=np.float32) / framerate ###时间横坐标
    depth = np.array([cal_depth(v0, times=i / 2) for i in t])

    data = data[:n] * (1+a*t)   #时变增益
    depth_sta_lta = '%.2f' % depth[np.argmax(cft[start:])+start]

    # plt.figure(figsize=(20, 10),facecolor='w')
    if len(depth) == len(redata):
        ax1 = plt.subplot(211)
        # plt.grid(True)
        # ax1.plot(depth,data, 'b',linewidth = 2,label="深度为：" + str(depth_sta_lta) + "米")  ###原始信号图
        # plt.xlabel("Depth")
        # plt.ylabel("Original Waveform")
        # plt.legend()
        ax1.vlines(x=depth[np.argmax(cft[start:])+start], ymin=np.min(data), ymax=np.max(data), linestyle='--',
                   linewidth=1.5,colors="k", alpha=1)
        ax1.hlines(y=data[np.argmax(cft[start:])+start], xmin=0, xmax=max(depth), linestyle='--', linewidth=1.5,
                   colors="k", alpha=1)

        ax2 = plt.subplot(212)
        # plt.grid(True)
        # ax2.plot(depth, cft, 'r',linewidth = 1)   ###sta/lta图
        # plt.xlabel("Depth")
        # plt.ylabel("STA/LTA")
        # plt.show()

    else:
        ax1 = plt.subplot(211)
        # plt.grid(True)
        # ax1.plot(depth, data[:], 'b', linewidth=2,label="深度为：" + str(depth_sta_lta) + "米")  ###原始信号图
        # plt.xlabel("depth")
        # plt.ylabel('Original Waveform')
        ax1.vlines(x=depth[np.argmax(cft[start:])+start], ymin=np.min(data), ymax=np.max(data), linestyle='--',
                   linewidth=1.5, colors="k", alpha=1)
        ax1.hlines(y=data[np.argmax(cft[start:])+start], xmin=0, xmax=max(depth), linestyle='--', linewidth=1.5,
                   colors="k", alpha=1)

        ax2 = plt.subplot(212)
        plt.grid(True)
        ax2.plot(depth[1:], cft[:], 'r', linewidth=1)  ###sta/lta图
        plt.xlabel("Depth")
        plt.ylabel("STA/LTA")
        plt.show()
    print("========================STA/LTA计算结果===========================")
    print('STA/LTA对应的索引', np.argmax(cft[start:]) + start)
    print('STA/LTA的深度为', depth_sta_lta, "m")
    print('STA/LTA的时间为', '%.2f' % (1000 * t[np.argmax(cft[start:]) + start]), "ms")
    print("\t")

# STA_LTA(channel1_data,nsta=10,nlta=100,a=0.2,predicted_depth=1800)



# def get_data(data):

#########################################################AIC##########################################
#预估深度 ： Predicted depth
def AIC(data,a,predicted_depth,flag):
    n = int(2 * predicted_depth * framerate / v0)
    data = data[:n]
    npts = len(data)  ##采样点个数
    start = int(2*(predicted_depth-800)*framerate/v0)
    ArgMax = np.argmax(data[start:])+start
    length = int((ArgMax-start)*0.75)

    # index_min = np.argmin(data[start:])+start-10000
    # index_max = index_min + length

    index_max = np.argmax(data[start:])+start
    index_min = index_max - length
    #
    #
    # index_min = 0
    # index_max = len(data)


    # index_max = np.argmax(data[index_min:])+index_min
    # print("长度",len(data[index_min:index_max]))
    # print(index_min,index_max)



    fc = np.square(data[index_min:index_max])  # 取平方

    # fc = []
    # for i in range(len(data[index_min:index_max])):
    #     data1 = data[i] - data[i-1]
    #     fc.append(data1)
    # fc = np.array(fc)

    # fc = []
    # for i in range(1,len(data[index_min:index_max])-1):
    #      data1 = np.square(data[i]) - (data[i-1] * data[i+1])
    #      fc.append(data1)
    # fc = np.array(fc)


    ############Kur-AIC算法(峰度AIC)##############
    # length = np.shape(fc)[0]
    # aic = []
    #
    #
    # for i in range(1, length +1):
    #     sum_fc = []
    #     sum_fc1 = []
    #     for j in range(1,i):
    #         aa = np.square(fc[j])
    #         sum_fc.append(aa)
    #     sum_fc = np.sum(sum_fc)
    #
    #     for m in range(i,length):
    #         bb = np.square(fc[m])
    #         sum_fc1.append(bb)
    #     sum_fc1 = np.sum(sum_fc1)
    #
    #     if i==0 or length-i-1 ==0:
    #         #print("被除数为0")
    #         continue
    #     aic_ = i * np.log((1/i)*sum_fc) + (length - i - 1) * np.log((1/(length - i - 1))*sum_fc1)
    #     aic.append(aic_)
    # aic = np.array(aic)


    ############VAR-AIC算法##############
    #length = np.shape(fc)[0]
    aic = []
    for i in range(1, length + 1):
        aic_ = i * np.log(np.var(fc[1:i])) + (length - i - 1) * np.log(np.var(fc[i+1:length]))
        aic.append(aic_)
    aic = np.array(aic)


    t1 = (np.arange(npts, dtype=np.float32) / framerate)[index_min:index_max] ###时间横坐标
    t2 = np.arange(npts, dtype=np.float32) / framerate  ###时间横坐标
    if flag == 0:
        depth1 = ((t1 * v0) / 2)
        depth2 = ((t2 * v0) / 2)
    else:
        depth1 = np.array([cal_depth(v0,times=i/2) for i in t1])
        depth2 = np.array([cal_depth(v0,times=i/2) for i in t2])


    # plt.figure(figsize=(20, 10),facecolor='w')
    ax1 = plt.subplot(211)
    # plt.grid(True)
    if len(depth1) == len(aic):
        ax1.plot(depth1[:], aic, 'r', label="aic")
        ax1.legend(loc=1,fontsize = 10)
        ax1.set_ylabel('AIC Values')
        ax1.set_xlabel('Depth')

        ax2 = ax1.twinx()  # this is the important function
    # ax2.plot(t, data, 'g', label="raw")
        ax2.plot(depth1, data[index_min:index_max], 'k', label="raw")
        ax2.legend(loc=2,fontsize=10)
        ax2.set_ylabel('Original Waveform')
    else:
        ax1.plot(depth1[1:], aic, 'r', label="aic")
        ax1.legend(loc=1, fontsize=10)
        ax1.set_ylabel('AIC Values')
        ax1.set_xlabel('Depth')

        ax2 = ax1.twinx()  # this is the important function
        # ax2.plot(t, data, 'g', label="raw")
        ax2.plot(depth1, data[index_min:index_max], 'k', label="raw")
        ax2.legend(loc=2, fontsize=10)
        ax2.set_ylabel('Original Waveform')


    ax3 = plt.subplot(212)
    # plt.grid(True)
    data = data * (1+a*t2)  # 时变增益
    depth_aic = '%.2f' % depth1[np.argmin(aic[10:-10]) + 10]
    ax3.plot(depth2,data,linewidth=2,label="深度为：" + str(depth_aic) + "米")
    # plt.xlabel("Depth")
    # plt.ylabel('Original Waveform')
    # for i in [index_min, index_max]:
    #     ax3.vlines(x=i,ymin=np.min(data),ymax=np.max(data),linestyle='--',linewidth=2,colors="r", alpha=1)
    ###在原数据中标出找到的最小值对应的点
    ax3.vlines(x=depth2[np.argmin(aic[10:-10])+10+index_min],ymin=np.min(data),ymax=np.max(data),linestyle='--',
               linewidth=1.5,colors="r", alpha=1)
    ax3.hlines(y=data[np.argmin(aic[10:-10])+10+index_min],xmin=0,xmax=max(depth2),linestyle='--',linewidth=1.5,
               colors="r", alpha=1)
    # plt.legend()
    # plt.show()
    # print("========================AIC计算结果=========================")
    print("由AIC算法预测的深度约为：", depth_aic, "m")
    # print(depth_aic)
    # print("AIC的时间为", '%.2f' % (1000 * t1[np.argmin(aic[10:-10]) + 10]), "ms")
    # print("AIC对应的索引", np.argmin(aic[10:-10]) + 10 + index_min)
    # print("index_min", index_min)
    # print("index_max", index_max)


# AIC(channel1_data,a=0.2,predicted_depth=1800,flag=3)

if __name__=='__main__':
     print("canshu=>",sys.argv[0],sys.argv[1])
    AIC(channel1_data, a=0.2, predicted_depth=(int(sys.argv[1])), flag=3)